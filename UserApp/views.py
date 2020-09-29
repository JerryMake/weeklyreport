import datetime
import time
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Q
from django.utils.http import urlquote
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render
from io import BytesIO
from openpyxl import Workbook
from UserApp import models
from UserApp.models import User, Report


def gologin(request):
    # 跳转登录页面
    return render(request,'login.html')

def login(request):
    # 登录逻辑
    if request.method == 'POST':
        userCode = request.POST.get('userCode')
        userPassword = request.POST.get('userPassword')
        if userCode == None and userPassword == None:
            result = '用户账号或密码不能为空！'
            return render(request,'login.html',{'result':result})
        else:
            # 进入数据库查询是否存在该用户账号,filter过滤函数,first()找出第一个
            login_user = User.objects.filter(userCode=userCode).first()
            if login_user:
                if userPassword == login_user.userPassword:
                    userName = login_user.userName
                    logo = login_user.logo
                    userCode = login_user.userCode
                    id = login_user.id
                    role = login_user.role
                    department = login_user.department
                    # 获取本周周数
                    this_week = time.strftime('%W')
                    return render(request,'index.html',{'userName':userName,'logo':logo,'userCode':userCode,'id':id,'role':role,'department':department,'this_week':this_week})
                else:
                    result = '密码错误！'
                    return render(request,'login.html',{'result':result})
            else:
                result = '账号不存在！'
                return render(request,'login.html',{'result':result})
    return render(request,'login.html')


def info(request):
    # 跳转用户信息页面
    return render(request,"info.html")

def go_changePassword(request):
    # 跳转修改用户密码页面
    return render(request,'changePassword.html')

def changePassword(request):
    # 修改用户密码
    if request.method=='POST':
        new_password1 = request.POST.get('new_password1')
        new_password2 = request.POST.get('new_password2')
        id = request.POST.get('id')
        if new_password1==new_password2:
            models.User.objects.filter(id=id).update(userPassword=new_password1)
            result = 'success'
            return render(request,'changePassword.html',{'result':result,'id':id})
        else:
            result = 'fail'
            return render(request,'changePassword.html',{'result':result,'id':id})

def add(request):
    # 跳转添加用户页面
    return render(request,'add.html')

def addUser(request):
    # 添加新用户
    if request.method=='POST':
        userCode = request.POST.get('userCode')
        userName = request.POST.get('userName')
        userPassword = request.POST.get('userPassword')
        logo = request.FILES.get('upload','')   # 获取上传图片
        role = request.POST.get('role')
        #保存图片
        user = User.objects.filter(userCode=userCode).first()
        if user:
            result = '该用户账号已存在！'
            return render(request,'add.html',{'result':result})
        else:
            models.User.objects.create(userCode=userCode,userName=userName,userPassword=userPassword,role=role,logo=logo,createTime=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            result = '添加成功！'
            return HttpResponseRedirect('/userList.html?result='+result)

def userList(request):
    # 所有用户信息
    if request.method == 'GET':
        # 获取当前页数（需要跳转的页数）
        num = request.GET.get('num',1)
        n = int(num)
        # 显示所有用户信息
        all_user = User.objects.all()
        if all_user.exists():
            paginator = Paginator(all_user,6)
            try:
                # 当前页数据
                paginator_data = paginator.page(n)
                # 返回跳转后的页数
                this_num = n
                # 返回最后一页页数
                end_num = int(paginator.count/6)+1
            except PageNotAnInteger:
                # 返回第一页数据
                paginator_data = paginator.page(1)
                # 返回跳转后的页数
                this_num = 1
                # 返回最后一页页数
                end_num = int(paginator.count/6)+1
            except EmptyPage:
                # 返回最后一页数据
                paginator_data = paginator.page(paginator.num_pages)
                # 返回跳转后的页数
                this_num = paginator.num_pages
                # 返回最后一页页数
                end_num = int(paginator.count/6)+1

            return render(request,'userList.html',{'paginator':paginator,'paginator_data':paginator_data,'this_num':this_num,'end_num':end_num})

def deleteUser(request):
    # 删除单条用户
    if request.method == "GET":
        delete_id = request.GET.get('id')
        models.User.objects.filter(id=delete_id).delete()
        return HttpResponseRedirect('/userList.html')

@csrf_exempt
def selectDelUser(request):
    # 批量删除选中的用户
    if request.method == "POST":
        id_list = request.POST.getlist('id_list')
        for i in id_list:
            if i!='':
                models.User.objects.filter(id=i).delete()
    return HttpResponseRedirect('/userList.html')

def modifyUser(request):
    # 修改用户信息
    if request.method == 'POST':
        id = request.POST.get('id')
        userName = request.POST.get('userName')
        userPassword = request.POST.get('userPassword')
        logo = request.FILES.get('upload')
        role = request.POST.get('role')
        department = request.POST.get('department')
        if logo != None:
            User.objects.filter(id=id).update(userName=userName,userPassword=userPassword,role=role,logo=logo,department=department)
            result = '修改成功'
            return  HttpResponseRedirect('/userList.html?result='+result)
        else:
            User.objects.filter(id=id).update(userName=userName,userPassword=userPassword,role=role)
            result = '修改成功'
            return HttpResponseRedirect('/userList.html?result='+result)

def goModify(request):
    # 进入修改用户
    # 找出需要修改的用户信息
    if request.method == 'GET':
        id = request.GET.get('id')
        modify_user = User.objects.filter(id=id).first()
        if modify_user:
            return render(request,'modifyUser.html',{'modify_user':modify_user})
    return ''

@csrf_exempt
def findUser(request):
    # 模糊查询
    if request.method == 'POST':
        # 获取当前页数（需要跳转的页数）
        num = request.GET.get('num',1)
        n = int(num)
        userName = request.POST.get('userName')
        role = request.POST.get('role')
        userCode = request.POST.get('userCode')
        print(userName)
        print(role)
        print(userCode)
        if userName!='' or role!='' or userCode !='':
            findUser = User.objects.filter(Q(userCode__contains=userCode) & Q(userName__contains=userName) & Q(role__contains=role))
            print(findUser)
            if findUser:
                paginator = Paginator(findUser,6)
                try:
                    # 当前页数据
                    paginator_data = paginator.page(n)
                    # 返回跳转后的页数
                    this_num = n
                    # 返回最后一页页数
                    end_num = int(paginator.count/6)+1
                except PageNotAnInteger:
                    # 返回第一页数据
                    paginator_data = paginator.page(1)
                    # 返回跳转后的页数
                    this_num = 1
                    # 返回最后一页页数
                    end_num = int(paginator.count/6)+1
                except EmptyPage:
                    # 返回最后一页数据
                    paginator_data = paginator.page(paginator.num_pages)
                    # 返回跳转后的页数
                    this_num = paginator.num_pages
                    # 返回最后一页页数
                    end_num = int(paginator.count/6)+1
                print(paginator_data)
                return render(request,'userList.html',{'paginator':paginator,'paginator_data':paginator_data,'this_num':this_num,'end_num':end_num})
    return render(request,'userList.html')

def addReport(request):
    if request.method == 'GET':
        userCode = request.GET.get('userCode')
        department = request.GET.get('department')
        # 获取本周周数
        this_week = time.strftime('%W')
        return render(request,'addReport.html',{'this_week':this_week,'userCode':userCode,'department':department})

def createReport(request):
    # 新增周报
    if request.method == 'POST':
        department = request.POST.get('department')
        reporter = request.POST.get('userCode')
        this_week_completed = request.POST.get('this_week_completed')
        this_week_not_completed = request.POST.get('this_week_not_completed')
        last_reson = request.POST.get('last_reson')
        next_week_plan = request.POST.get('next_week_plan')
        need_help = request.POST.get('need_help')
        week_number = request.POST.get('week_number')
        if department!='' and reporter!='':
            findUser = User.objects.filter(userCode=reporter).first()
            print(findUser)
            reporter_name = findUser.userName
            print(reporter_name)
            models.Report.objects.create(week_number=week_number,department=department,reporter=reporter,this_week_completed=this_week_completed,
                                         this_week_not_completed=this_week_not_completed,last_reson=last_reson,next_week_plan=next_week_plan,
                                         need_help=need_help,reporter_name=reporter_name)
        else:
            return HttpResponse('部门不能为空')
        return HttpResponse('周报填写成功！')

def show_myReport(request):
    # 进入我的周报
    if request.method == 'GET':
        userCode = request.GET.get('userCode')
        department = request.GET.get('department')
        if userCode !='':
            findReport = Report.objects.filter(reporter=userCode).all().order_by('-week_number')
            return render(request,'myReport.html',{'findReport':findReport,'department':department,'userCode':userCode})
        else:
            return HttpResponse('查询不到周报！')

    elif request.method == 'POST':
        reporter = request.POST.get('reporter')
        week_number = request.POST.get('week_number')
        department = request.POST.get('department')
        if week_number!=''and reporter!='':
            findReport = Report.objects.filter(Q(week_number=week_number)&Q(reporter=reporter)).all()
            return render(request,'myReport.html',{'findReport':findReport,'userCode':reporter,'department':department})
        else:
            return HttpResponse('查询不到数据！')

def thisWeek(request):
    # 进入本周周报
    if request.method == 'GET':
        week_number = request.GET.get('this_week')
        if week_number!='':
            this_week = Report.objects.filter(Q(week_number=week_number)).all()
            return render(request,'thisWeek.html',{'this_week':this_week})

    if request.method == 'POST':
        # 导出excel
        # 获取本周周数
        this_week = time.strftime('%W')
        wb = Workbook()		# 生成一个工作簿（即一个Excel文件）
        wb.encoding = 'utf-8'
        sheet1 = wb.active	# 获取第一个工作表（sheet1）
        sheet1.title = '挂号信息'	# 给工作表1设置标题
        row_one = [ '名称', '部门', '本周完成的计划工作', '本周未完成的计划工作','未完成原因', '下周计划', '需要公司协助事项', ]
        for i in range(1, len(row_one)+1):	# 从第一行开始写，因为Excel文件的行号是从1开始，列号也是从1开始
            # 从row=1，column=1开始写，即将row_one的数据依次写入第一行
            sheet1.cell(row=1, column=i).value=row_one[i-1]
            all_obj = Report.objects.filter(week_number=this_week).all()
        for obj in all_obj:
            max_row = sheet1.max_row + 1	# 获取到工作表的最大行数并加1
            obj_info = [obj.reporter_name,obj.department,obj.this_week_completed,obj.this_week_not_completed,obj.last_reson,obj.next_week_plan,obj.need_help]
            for x in range(1, len(obj_info)+1):		# 将每一个对象的所有字段的信息写入一行内
                sheet1.cell(row=max_row, column=x).value = obj_info[x-1]

            # 准备写入到IO中
        output = BytesIO()
        wb.save(output)	 # 将Excel文件内容保存到IO中
        output.seek(0)	 # 重新定位到开始
        # 设置HttpResponse的类型
        response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
        ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        file_name = '周报%s.xls' % ctime	# 给文件名中添加日期时间
        file_name = urlquote(file_name)	 # 使用urlquote()方法解决中文无法使用的问题
        response['Content-Disposition'] = 'attachment; filename=%s' % file_name
        # response.write(output.getvalue())	 # 在设置HttpResponse的类型时，如果给了值，可以不写这句
        return response

